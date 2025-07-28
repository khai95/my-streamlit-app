import os
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from streamlit_autorefresh import st_autorefresh

st.set_page_config(page_title="Rain Odds & Chart", layout="wide")
st.title("Real-time Rain Odds and Chart")

# 手动输入 Excel 文件名（Sidebar）
filename_input = st.sidebar.text_input(
    "请输入 Excel 文件名 (如 JUL28_06-18.xlsx):",
    value="JUL28_06-18.xlsx"
)

# 完整的 Excel 文件路径
excel_path = os.path.join(
    os.path.expanduser('~'),
    'Desktop',
    'Khai',
    'krain',
    'krain99',
    'Record',
    filename_input
)

# 自动每 N 秒刷新（Sidebar可调整）
refresh_interval = st.sidebar.slider("Refresh interval (seconds)", 2, 30, 2)
st_autorefresh(interval=refresh_interval*1000, key="datarefresh")

def read_data():
    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                data.append({
                    "Rain Odds": float(row[0]),
                    "50MA": float(row[1]) if row[1] is not None else None,
                    "150MA": float(row[2]) if row[2] is not None else None,
                    "Time": row[3]
                })
        return pd.DataFrame(data)
    except FileNotFoundError:
        st.error(f"找不到文件：{filename_input}，请确认输入是否正确。")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"发生错误：{e}")
        return pd.DataFrame()

df = read_data()

if not df.empty:
    st.line_chart(df[["Rain Odds", "50MA", "150MA"]])
    st.dataframe(df)
    st.info(f"資料最新時間：{df['Time'].iloc[-1]}")
else:
    st.warning("尚無資料，請確認 Excel 檔案名稱正確，且監控腳本已開始寫入 Excel。")
